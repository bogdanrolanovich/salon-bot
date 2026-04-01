import asyncio
import datetime
import os
from aiogram import Bot, Dispatcher, types
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from aiogram.filters import Command
from openpyxl import Workbook, load_workbook

# ===== НАСТРОЙКИ =====
import os

TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = 778497367  # <-- вставь свой Telegram ID

services = ["Стрижка", "Маникюр", "Педикюр", "Окрашивание"]
time_slots = ["10:00", "11:00", "12:00", "13:00", "14:00", "15:00"]

excel_file = "appointments.xlsx"

# ===== СОЗДАНИЕ EXCEL =====
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Клиент", "Услуга", "Дата", "Время"])
    wb.save(excel_file)

# ===== СОСТОЯНИЕ ПОЛЬЗОВАТЕЛЕЙ =====
user_state = {}

# ===== ПРОВЕРКА ЗАНЯТОСТИ =====
def is_time_taken(date, time):
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == date and row[3] == time:
            return True
    return False

# ===== ЗАПУСК БОТА =====
async def main():
    bot = Bot(token=TOKEN)
    dp = Dispatcher()

    # ===== /start =====
    @dp.message(Command("start"))
    async def start(message: types.Message):
        name = message.from_user.first_name

        text = (
            f"💎 Привет, {name}! Добро пожаловать в салон красоты!\n\n"
            "Мы поможем вам выглядеть идеально ✨\n\n"
            "💇‍♀️ Наши услуги:\n"
            "• Стрижка\n"
            "• Маникюр\n"
            "• Педикюр\n"
            "• Окрашивание\n\n"
            "📅 Запись занимает всего 1 минуту!\n\n"
            "👇 Выберите услугу:"
        )

        buttons = [KeyboardButton(text=s) for s in services]
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[b] for b in buttons],
            resize_keyboard=True
        )

        await message.answer(text, reply_markup=keyboard)
        user_state[message.from_user.id] = {"step": "service"}

    # ===== ОСНОВНАЯ ЛОГИКА =====
    @dp.message()
    async def handle(message: types.Message):
        user_id = message.from_user.id

        if user_id not in user_state:
            await message.answer("Напишите /start чтобы начать запись.")
            return

        state = user_state[user_id]

        # --- 1. ВЫБОР УСЛУГИ ---
        if state["step"] == "service" and message.text in services:
            state["service"] = message.text
            state["step"] = "date"

            today = datetime.date.today()
            dates = [(today + datetime.timedelta(days=i)).strftime("%d.%m.%Y") for i in range(3)]

            keyboard = ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text=d)] for d in dates],
                resize_keyboard=True
            )

            await message.answer("📅 Выберите дату:", reply_markup=keyboard)
            return

        # --- 2. ВЫБОР ДАТЫ ---
        if state["step"] == "date":
            state["date"] = message.text
            state["step"] = "time"

            keyboard = ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text=t)] for t in time_slots],
                resize_keyboard=True
            )

            await message.answer("⏰ Выберите время:", reply_markup=keyboard)
            return

        # --- 3. ВЫБОР ВРЕМЕНИ ---
        if state["step"] == "time" and message.text in time_slots:
            date = state["date"]
            time = message.text

            # 🔴 ПРОВЕРКА ЗАНЯТОСТИ
            if is_time_taken(date, time):
                await message.answer("❌ Это время уже занято, выберите другое.")
                return

            # Сохранение в Excel
            wb = load_workbook(excel_file)
            ws = wb.active
            ws.append([message.from_user.full_name, state["service"], date, time])
            wb.save(excel_file)

            # Клиенту
            await message.answer(
                f"✅ Запись подтверждена!\n\n"
                f"💇 Услуга: {state['service']}\n"
                f"📅 Дата: {date}\n"
                f"⏰ Время: {time}\n\n"
                f"Ждём вас! 💖"
            )

            # Администратору
            await bot.send_message(
                ADMIN_ID,
                f"🔥 Новая запись!\n\n"
                f"👤 Клиент: {message.from_user.full_name}\n"
                f"💇 Услуга: {state['service']}\n"
                f"📅 Дата: {date}\n"
                f"⏰ Время: {time}"
            )

            del user_state[user_id]
            return

        await message.answer("Пожалуйста, используйте кнопки ниже 👇")

    print("🚀 Бот запущен...")

    # ===== СТАБИЛЬНЫЙ ЗАПУСК =====
    while True:
        try:
            await dp.start_polling(bot)
        except Exception as e:
            print(f"Ошибка: {e}")
            await asyncio.sleep(5)

# ===== СТАРТ =====
if __name__ == "__main__":
    asyncio.run(main())